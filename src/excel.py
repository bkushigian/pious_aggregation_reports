from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import coordinate_to_tuple, get_column_letter
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Font, NamedStyle

import pandas as pd
from argparse import ArgumentParser
import os
from os import path as osp


def freeze_row(ws: Worksheet, row):
    if ws.freeze_panes is None:
        r, c = coordinate_to_tuple("A1")
    else:
        r, c = coordinate_to_tuple(str(ws.freeze_panes))
    c = " ABCDEFGHIJKLMNOPQRSTUVWXYZ"[c]
    ws.freeze_panes = f"{c}{row}"


def freeze_column(ws: Worksheet, column: int):
    if ws.freeze_panes is None:
        r, c = coordinate_to_tuple("A1")
    else:
        r, c = coordinate_to_tuple(str(ws.freeze_panes))
    column = " ABCDEFGHIJKLMNOPQRSTUVWXYZ"[column]
    ws.freeze_panes = f"{column}{r}"


def add_headers(df: pd.DataFrame, ws: Worksheet):
    # Write headers
    split_columns = [col.split(":") for col in df.columns]
    max_depth = max(len(split) for split in split_columns)
    col_idx = 1

    col_types = {"raise": [], "bet": [], "check": [], "fold": [], "call": [], "ev": []}

    for i, split in enumerate(split_columns, start=1):
        row_idx = 1
        found = False
        for sub_col in split:
            sc = sub_col.strip().lower()
            if sc in col_types:
                if found:
                    raise RuntimeError("Illegal State")
                found = True
                col_types[sc].append(i)
            elif sc in ("oop ev", "ip_ev"):
                if found:
                    raise RuntimeError("Illegal State")
                found = True
                col_types["ev"].append(i)
            cell = ws.cell(row=row_idx, column=col_idx, value=sub_col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            row_idx += 1
        col_idx += 1
    return max_depth, col_types


def format_card_columns(ws: Worksheet) -> int:
    column_names = ("flop", "turn", "river")
    suit_symbols = {"h": "♥", "d": "♦", "s": "♠", "c": "♣"}
    suit_colors = {"h": "FF0000", "d": "0000FF", "s": "000000", "c": "00FF00"}
    max_card_col_index = 0
    for col_idx in range(1, 4):
        cell = ws.cell(row=1, column=col_idx)
        if str(cell.value).lower() in column_names:
            max_card_col_index += 1
            row_idx = 2
            while row_idx <= ws.max_row:
                cell = ws.cell(row=row_idx, column=col_idx)
                v = cell.value
                if v is not None:
                    v = str(cell.value)
                    xs = []
                    for i in range(0, len(v), 2):
                        rank = v[i]
                        suit = v[i + 1]
                        sym = suit_symbols.get(suit, "?")
                        color = suit_colors.get(suit, "000000")
                        card_text = f"{rank}{sym}"
                        xs.append(card_text)

                    joined = " ".join(xs)
                    cell = ws.cell(row=row_idx, column=col_idx, value=joined)
                row_idx += 1

    # Now, freeze these columns
    freeze_column(ws, max_card_col_index + 1)
    return max_card_col_index


def format_cells_as_bars(ws: Worksheet, col_types, header_height):
    row_idx = header_height

    raise_databar_rule = DataBarRule(
        start_type="num",
        start_value=0,
        end_type="num",
        end_value=100,
        color="ad1616",
    )
    call_databar_rule = DataBarRule(
        start_type="num",
        start_value=0,
        end_type="num",
        end_value=100,
        color="00a933",
    )

    bet_databar_rule = DataBarRule(
        start_type="num",
        start_value=0,
        end_type="num",
        end_value=100,
        color="bd1010",
    )
    check_databar_rule = DataBarRule(
        start_type="num",
        start_value=0,
        end_type="num",
        end_value=100,
        color="00c922",
    )
    fold_databar_rule = DataBarRule(
        start_type="num",
        start_value=0,
        end_type="num",
        end_value=100,
        color="0033c9",
    )
    for raise_col in col_types["raise"]:
        c = get_column_letter(raise_col)
        r = header_height + 1
        top_left = f"{c}{r}"
        bot_right = f"{c}{ws.max_row}"
        ws.conditional_formatting.add(f"{top_left}:{bot_right}", raise_databar_rule)
    for call_col in col_types["call"]:
        c = get_column_letter(call_col)
        r = header_height + 1
        top_left = f"{c}{r}"
        bot_right = f"{c}{ws.max_row}"
        ws.conditional_formatting.add(f"{top_left}:{bot_right}", call_databar_rule)
    for col in col_types["bet"]:
        c = get_column_letter(col)
        r = header_height + 1
        top_left = f"{c}{r}"
        bot_right = f"{c}{ws.max_row}"
        ws.conditional_formatting.add(f"{top_left}:{bot_right}", bet_databar_rule)
    for col in col_types["check"]:
        c = get_column_letter(col)
        r = header_height + 1
        top_left = f"{c}{r}"
        bot_right = f"{c}{ws.max_row}"
        ws.conditional_formatting.add(f"{top_left}:{bot_right}", check_databar_rule)
    for col in col_types["fold"]:
        c = get_column_letter(col)
        r = header_height + 1
        top_left = f"{c}{r}"
        bot_right = f"{c}{ws.max_row}"
        ws.conditional_formatting.add(f"{top_left}:{bot_right}", fold_databar_rule)


def process_df(df: pd.DataFrame, ws: Worksheet) -> int:
    n_columns = len(df.columns)

    max_depth, col_types = add_headers(df, ws)

    # Now merge, starting at the top
    def recursively_merge_headers(row_idx, col_idx, max_columns=None):
        if max_columns is None:
            max_columns = n_columns
        if row_idx > max_depth:
            return col_idx
        while col_idx < max_columns:
            cell = ws.cell(row=row_idx, column=col_idx)
            value = cell.value
            v = value
            right = col_idx
            while right < n_columns and v == value and str(value) != "":
                right += 1
                cell = ws.cell(row=row_idx, column=right)
                v = cell.value

            if right > col_idx + 1:
                ws.merge_cells(
                    start_row=row_idx,
                    start_column=col_idx,
                    end_row=row_idx,
                    end_column=right - 1,
                )
                # Now recursively merge subcolumns
                recursively_merge_headers(row_idx + 1, col_idx, right)
            col_idx = right

    recursively_merge_headers(row_idx=1, col_idx=1)

    freeze_row(ws, max_depth + 1)
    return max_depth, col_types


def main():
    parser = ArgumentParser()
    parser.add_argument("dir", help="directory to turn into workbook")
    parser.add_argument(
        "--out", help="filename to save to", default="MyAggregationWorkbook.xlsx"
    )

    args = parser.parse_args()
    dir = args.dir

    files = os.listdir(dir)

    wb = Workbook()
    ws = None
    one_decimal_style = NamedStyle(name="one_decimal_style", number_format="0.0")
    file_order = [
        "Aggregation",
        "Overview",
        "Nothing",
        "Pairs",
        "OverPair",
        "TopPair",
        "UnderPair(1)",
        "2ndPair",
        "UnderPair(2)",
        "3rdPair",
        "UnderPair(3)",
        "4thPair",
        "UnderPair(4)",
        "5thPair",
        "UnderPair(5)",
    ]
    unvisited = {f for f in files if f.endswith(".csv")}
    for file_base in file_order:
        file = file_base + ".csv"
        if file not in unvisited:
            continue
        unvisited.remove(file)

        rel_path = osp.join(dir, file)
        path = osp.abspath(rel_path)
        try:
            df = pd.read_csv(path)
        except pd.errors.EmptyDataError:
            continue
        title = file[:-4]
        if ws is None:
            ws = wb.active
            ws.title = title
        else:
            ws = wb.create_sheet(title=title)

        max_depth, col_types = process_df(df, ws)
        for r_idx, row in enumerate(
            dataframe_to_rows(df, index=False, header=False), start=max_depth + 1
        ):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.style = one_decimal_style
        format_card_columns(ws)
        format_cells_as_bars(ws, col_types, max_depth)

    if len(unvisited) > 0:
        print(f"Warning: unprocessed CSVs: {unvisited}")
    out = args.out
    if not out.endswith(".xlsx"):
        out = f"{out}.xlsx"
    wb.save(out)
    print("Dataframes added to workbook!")


if __name__ == "__main__":
    main()
