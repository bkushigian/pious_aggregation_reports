from typing import Dict
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import coordinate_to_tuple, get_column_letter
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Font, NamedStyle

import pandas as pd
from argparse import ArgumentParser
import os
from os import path as osp
from ansi.color import fg
from tqdm import tqdm
from .trial import *

CARD_COLUMN_WIDTH = 3.75
ACTION_COLUMN_WIDTH = 8


def freeze_row(ws: Worksheet, row):
    if ws.freeze_panes is None:
        r, c = coordinate_to_tuple("A1")
    else:
        r, c = coordinate_to_tuple(str(ws.freeze_panes))
    c = " ABCDEFGHIJKLMNOPQRSTUVWXYZ"[c]
    ws.freeze_panes = f"{c}{row}"


def freeze_column(ws: Worksheet, column: int):
    if ws.freeze_panes is None:
        r, c = coordinate_to_tuple("A1")
    else:
        r, c = coordinate_to_tuple(str(ws.freeze_panes))
    column = " ABCDEFGHIJKLMNOPQRSTUVWXYZ"[column]
    ws.freeze_panes = f"{column}{r}"


def add_headers(df: pd.DataFrame, ws: Worksheet):
    # Write headers
    split_columns = [col.split(":") for col in df.columns]
    max_depth = max(len(split) for split in split_columns)
    col_idx = 1

    col_types = {"raise": [], "bet": [], "check": [], "fold": [], "call": [], "ev": []}
    actions = {"raise", "bet", "check", "fold", "call"}

    for i, split in enumerate(split_columns, start=1):
        row_idx = 1
        found = False
        for sub_col in split:
            sc = sub_col.strip().lower()
            if sc in col_types:
                if found:
                    raise RuntimeError("Illegal State")
                found = True
                col_types[sc].append(i)

                if sc in actions:
                    col_letter = get_column_letter(col_idx)
                    ws.column_dimensions[col_letter].width = ACTION_COLUMN_WIDTH
            elif sc in ("oop ev", "ip_ev"):
                if found:
                    raise RuntimeError("Illegal State")
                found = True
                col_types["ev"].append(i)
            cell = ws.cell(row=row_idx, column=col_idx, value=sub_col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            row_idx += 1
        col_idx += 1
    return max_depth, col_types


def format_card_columns(ws: Worksheet) -> int:
    column_names = ("flop", "turn", "river")
    suit_symbols = {"h": "♥", "d": "♦", "s": "♠", "c": "♣"}
    suit_colors = {
        "h": Font(color="AA1200"),
        "d": Font(color="0000FF"),
        "s": Font(color="000000"),
        "c": Font(color="00AA23"),
    }
    card_map = {}
    for r in "23456789TJQKA":
        for s in "hdsc":
            card = f"{r}{suit_symbols[s]}"
            font = suit_colors[s]
            card_map[f"{r}{s}"] = (card, font)
    max_card_col_index = 0
    for col_idx, col in enumerate(
        ws.iter_cols(min_row=2, max_row=ws.max_row, max_col=5), start=1
    ):
        cell = ws.cell(row=2, column=col_idx)
        if str(cell.value).lower() in column_names:
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = CARD_COLUMN_WIDTH
            max_card_col_index += 1

            for row_idx, cell in enumerate(col[2:], start=4):
                if row_idx % 100 == 0:
                    print(
                        f"\r      {col_idx}, {row_idx}/{ws.max_row} ({100 *row_idx/ws.max_row:5.1f}%)",
                        end="",
                    )
                v = cell.value
                if v is not None:
                    card, font = card_map[v]
                    cell = ws.cell(row=row_idx, column=col_idx, value=card)
                    cell.font = font

    # Now, freeze these columns
    freeze_column(ws, max_card_col_index + 1)
    return max_card_col_index


def format_cells_as_bars(ws: Worksheet, col_types, header_height):
    row_idx = header_height

    raise_databar_rule = DataBarRule(
        start_type="num",
        start_value=0,
        end_type="num",
        end_value=100,
        color="ad1616",
    )
    call_databar_rule = DataBarRule(
        start_type="num",
        start_value=0,
        end_type="num",
        end_value=100,
        color="00a933",
    )

    bet_databar_rule = DataBarRule(
        start_type="num",
        start_value=0,
        end_type="num",
        end_value=100,
        color="bd1010",
    )
    check_databar_rule = DataBarRule(
        start_type="num",
        start_value=0,
        end_type="num",
        end_value=100,
        color="00c922",
    )
    fold_databar_rule = DataBarRule(
        start_type="num",
        start_value=0,
        end_type="num",
        end_value=100,
        color="0033c9",
    )
    for raise_col in col_types["raise"]:
        c = get_column_letter(raise_col)
        r = header_height + 1
        top_left = f"{c}{r}"
        bot_right = f"{c}{ws.max_row}"
        ws.conditional_formatting.add(f"{top_left}:{bot_right}", raise_databar_rule)
    for call_col in col_types["call"]:
        c = get_column_letter(call_col)
        r = header_height + 1
        top_left = f"{c}{r}"
        bot_right = f"{c}{ws.max_row}"
        ws.conditional_formatting.add(f"{top_left}:{bot_right}", call_databar_rule)
    for col in col_types["bet"]:
        c = get_column_letter(col)
        r = header_height + 1
        top_left = f"{c}{r}"
        bot_right = f"{c}{ws.max_row}"
        ws.conditional_formatting.add(f"{top_left}:{bot_right}", bet_databar_rule)
    for col in col_types["check"]:
        c = get_column_letter(col)
        r = header_height + 1
        top_left = f"{c}{r}"
        bot_right = f"{c}{ws.max_row}"
        ws.conditional_formatting.add(f"{top_left}:{bot_right}", check_databar_rule)
    for col in col_types["fold"]:
        c = get_column_letter(col)
        r = header_height + 1
        top_left = f"{c}{r}"
        bot_right = f"{c}{ws.max_row}"
        ws.conditional_formatting.add(f"{top_left}:{bot_right}", fold_databar_rule)


def process_card_columns(df: pd.DataFrame) -> pd.DataFrame:
    flops = df["Flop"]
    fc1 = []
    fc2 = []
    fc3 = []
    new_columns = {"Board:Flop:1": fc1, "Board:Flop:2": fc2, "Board:Flop:3": fc3}
    to_drop = ["Flop"]
    for flop in flops:
        c1 = flop[0:2]
        c2 = flop[2:4]
        c3 = flop[4:6]
        fc1.append(c1)
        fc2.append(c2)
        fc3.append(c3)
    if "Turn" in df:
        turns = df["Turn"]
        new_columns["Board:Turn"] = turns
        to_drop.append("Turn")

    if "River" in df:
        rivers = df["River"]
        new_columns["Board:River"] = rivers
        to_drop.append("River")
    return pd.concat([pd.DataFrame(data=new_columns), df.drop(columns=to_drop)], axis=1)


def process_df(df: pd.DataFrame, ws: Worksheet) -> int:

    max_depth, col_types = add_headers(df, ws)

    freeze_row(ws, max_depth + 1)
    return max_depth, col_types


def make_directory_for_path(path):
    d = osp.dirname(path)
    print("dir", d)
    os.makedirs(d)


def main():
    parser = ArgumentParser()
    parser.add_argument("dir", help="directory to turn into workbook")
    parser.add_argument(
        "--out", help="filename to save to", default="MyAggregationWorkbook.xlsx"
    )

    args = parser.parse_args()
    wb = make_workbook_from_dir(args.dir)
    out = args.out
    if not out.endswith(".xlsx"):
        out = f"{out}.xlsx"
    make_directory_for_path(out)
    wb.save(out)
    print(fg.green("Dataframes added to workbook!"))


WB_SHEET_ORDER = [
    "Aggregation",
    "Overview",
    "Nothing",
    "Pairs",
    "OverPair",
    "TopPair",
    "UnderPair(1)",
    "2ndPair",
    "UnderPair(2)",
    "3rdPair",
    "UnderPair(3)",
    "4thPair",
    "UnderPair(4)",
    "5thPair",
    "UnderPair(5)",
]


def make_workbook_from_dict(sub_sheets_map: Dict[str, pd.DatetimeIndex]) -> Workbook:
    print(fg.green(f"Making workbook from dict"))
    one_decimal_style = NamedStyle(name="one_decimal_style", number_format="0.0")
    wb = Workbook()
    ws = None
    unvisited_keys = set(sub_sheets_map.keys())
    for title in WB_SHEET_ORDER:
        if title not in sub_sheets_map:
            print(fg.yellow(f"   No CSV for {title}"))
            continue
        unvisited_keys.remove(title)
        df = sub_sheets_map[title]
        if ws is None:
            ws = wb.active
            ws.title = title
        else:
            ws = wb.create_sheet(title=title)

        df = process_card_columns(df)
        max_depth, col_types = process_df(df, ws)
        rows = list(dataframe_to_rows(df, index=False, header=False))
        for r_idx, row in tqdm(
            enumerate(rows, start=max_depth + 1),
            total=len(rows),
            desc=fg.boldcyan(f"{title:13}"),
        ):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.style = one_decimal_style

        # Now merge, starting at the top
        def recursively_merge_headers(row_idx, col_idx, max_columns=None):
            if max_columns is None:
                max_columns = n_columns
            if row_idx > max_depth:
                return col_idx
            while col_idx < max_columns:
                cell = ws.cell(row=row_idx, column=col_idx)
                value = cell.value
                v = value
                right = col_idx
                while right <= n_columns and v == value and str(value) != "":
                    right += 1
                    cell = ws.cell(row=row_idx, column=right)
                    v = cell.value

                if right > col_idx + 1:
                    ws.merge_cells(
                        start_row=row_idx,
                        start_column=col_idx,
                        end_row=row_idx,
                        end_column=right - 1,
                    )
                    # Now recursively merge subcolumns
                    recursively_merge_headers(row_idx + 1, col_idx, right)
                col_idx = right

        n_columns = len(df.columns)

        # print("   - Formatting Card Columns")
        format_card_columns(ws)
        # print("   - Formatting Cells as Bars")
        format_cells_as_bars(ws, col_types, max_depth)
        # print("   - Merging Headers")
        recursively_merge_headers(row_idx=1, col_idx=1)
    if len(unvisited_keys) > 0:
        print(f"Warning: unprocessed subsheets: {unvisited_keys}")

    return wb


def make_workbook_from_dir(dir: str):
    print(fg.green(f"Making workbook from {dir}\033[0m"))
    files = os.listdir(dir)
    unvisited_files = {f for f in files if f.endswith(".csv")}

    d = {}
    for file_base in WB_SHEET_ORDER:
        file = file_base + ".csv"
        if file not in unvisited_files:
            continue
        unvisited_files.remove(file)

        rel_path = osp.join(dir, file)
        path = osp.abspath(rel_path)
        try:
            df = pd.read_csv(path)
        except pd.errors.EmptyDataError:
            continue
        if len(df.columns) <= 1:
            continue
        title = file[:-4]
        d[title] = df
    if len(unvisited_files) > 0:
        print(f"{fg.yellow(Warning)}: unprocessed CSVs: {unvisited_files}")
    return make_workbook_from_dict(d)


if __name__ == "__main__":
    main()
